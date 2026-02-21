from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from io import BytesIO

class Stylizer:
    @staticmethod
    def style_excel(input_buffer: BytesIO) -> BytesIO:
        """
        Applies header style, zebra striping, and thin borders to the Excel file (in-memory).
        Args:
            input_buffer (BytesIO): Input Excel file in memory.
        Returns:
            BytesIO: Styled Excel file in memory.
        """
        input_buffer.seek(0)
        wb = load_workbook(input_buffer)
        ws = wb.active

        # Header style
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        # Apply header style
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border

        # Zebra striping and borders
        zebra_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border
                if (cell.row % 2) == 0:
                    cell.fill = zebra_fill

        output_buffer = BytesIO()
        wb.save(output_buffer)
        output_buffer.seek(0)
        wb.close()
        return output_buffer
