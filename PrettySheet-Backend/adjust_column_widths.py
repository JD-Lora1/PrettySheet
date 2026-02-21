#!/usr/bin/env python3
"""
adjust_column_widths.py

A small utility that adjusts Excel column widths based on cell content.

Usage:
  python adjust_column_widths.py path/to/file.xlsx
  python adjust_column_widths.py path/to/file.xlsx --sheet Sheet1 --min-width 8 --max-width 60 --padding 2

The script opens the workbook with openpyxl (preserving styles), computes
the maximum string length in each column and sets an approximate width.
"""

import argparse
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def _cell_length(value):
    if value is None:
        return 0
    # Convert to string and measure. This is a simple approximation.
    s = str(value)
    return len(s)


def adjust_widths(file_name, sheet=None, min_width=80, max_width=160, padding=2, out_file=None):
    wb = load_workbook(file_name)
    ws = wb[sheet] if sheet else wb.active

    max_lengths = {}
    for row in ws.iter_rows(values_only=True):
        for idx, cell in enumerate(row, start=1):
            cur = max_lengths.get(idx, 0)
            l = _cell_length(cell)
            if l > cur:
                max_lengths[idx] = l

    for idx, length in max_lengths.items():
        col = get_column_letter(idx)
        width = length + padding
        if width < min_width:
            width = min_width
        if width > max_width:
            width = max_width
        ws.column_dimensions[col].width = width

    target = out_file if out_file else file_name
    wb.save(target)
    return target


def main():
    parser = argparse.ArgumentParser(description="Adjust Excel column widths based on content")
    parser.add_argument("file", help="Path to the Excel file (.xlsx)")
    parser.add_argument("--sheet", help="Sheet name (default: active)")
    parser.add_argument("--min-width", type=int, default=80, help="Minimum column width (chars)")
    parser.add_argument("--max-width", type=int, default=160, help="Maximum column width (chars)")
    parser.add_argument("--padding", type=int, default=2, help="Extra chars to add to measured length")
    parser.add_argument("--out", help="Output file (if omitted, overwrites input)")

    args = parser.parse_args()
    out = adjust_widths(
        args.file,
        sheet=args.sheet,
        min_width=args.min_width,
        max_width=args.max_width,
        padding=args.padding,
        out_file=args.out,
    )
    print(f"Saved: {out}")


if __name__ == "__main__":
    main()
