from openpyxl import load_workbook
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from io import BytesIO

class Stylizer:
    @staticmethod
    def style_excel(input_buffer, column_configs, txt_color):
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
        
        wb = openpyxl.load_workbook(input_buffer)
        ws = wb.active
        
        # Limpieza del color de texto
        txt_color = str(txt_color).replace("#", "")
        header_font = Font(color=txt_color, bold=True, size=12)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                            top=Side(style='thin'), bottom=Side(style='thin'))

        # Iterar sobre las celdas del encabezado
        for cell in ws[1]:
            col_name = str(cell.value).strip() # Limpiamos espacios
            
            # Buscamos el color en el diccionario
            if col_name in column_configs:
                bg_hex = column_configs[col_name]["bg_color"].replace("#", "")
                cell.fill = PatternFill(start_color=bg_hex, end_color=bg_hex, fill_type="solid")
            else:
                # Color gris por defecto si no lo encuentra
                cell.fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
            
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        # Zebra striping and borders
        zebra_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border
                if (cell.row % 2) == 0:
                    cell.fill = zebra_fill

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
